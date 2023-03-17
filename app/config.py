"""
設定情報スプレッドシートからデータを取得したり算出したりするモジュール
"""

from app.model import Product, ProductInput
from dataclasses import dataclass
from datetime import datetime
from logging import getLogger
from more_itertools import first_true
from openpyxl.cell.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from typing import Any, Callable, Dict, List, Mapping, NamedTuple, Optional, Tuple, Union, cast
import itertools
import more_itertools
import re

logger = getLogger(__name__)

before_ext = re.compile(r'(?=\.(docx|xlsx))')


# DEEDの様式番号
form_no_deed = {11, 12, 21}

# NOTEの様式番号
form_no_note = {13, 14, 20}

ConfigValue = Union[str, int, float, datetime]


@dataclass(frozen=True)
class PropertyInfo:
    name: str  # 物件名
    keywords: Mapping[str, Any]  # 商品ごとのキーワードと値のペア
    jikkin_sheet: Optional[Workbook] = None


class RepByStateRecord(NamedTuple):
    form_no: int  # 様式番号
    form_name: str  # 様式名 (使わないが、excelとの対応を取るために追加)
    key: str  # 項目名 (のちに作るkey value の key)
    state: str  # 州名
    value: Any  # 項目の内容 (のちに作るkey value の value)


class RepByProductRecord(NamedTuple):
    name: str  # 商品区分
    form_no: int  # 様式番号
    form_name: str  # 様式名 (使わないが、excelとの対応を取るために追加)
    key: str  # 項目名 (のちに作るkey value の key)
    value: Any  # 項目の内容 (のちに作るkey value の value)


def cell_normalize(cell: Cell) -> ConfigValue:

    value = cell.value

    if isinstance(value, float):
        return int(value) if value % 1 == 0 else value
    else:
        return value


class Config:
    """
    設定情報.xlsxが持つような情報を抽象化したもの。
    """

    def __init__(self,
                 # 商品区分, 様式番号, 様式名, テンプレートファイル名
                 mapping: List[Tuple[ConfigValue, ...]],

                 keywords_rep_by_states: List[RepByStateRecord],

                 keywords_rep_by_product: List[RepByProductRecord],

                 # 項目名, 内容
                 keywords_global: List[Tuple[str, ...]]):

        self.__mapping = mapping
        self.__keywords_rep_by_state = keywords_rep_by_states
        self.__keywords_rep_by_product = keywords_rep_by_product
        self.__global = keywords_global

    def __specialize_filename(self, template_filename: str, form_no: int, state: str, fiance: bool) -> str:
        """
        設定情報から取得したテンプレートファイル名を州ごとに別名に変更するなどの特殊化を行う
        """

        # DEEDに対しての特殊処理
        if form_no in form_no_deed:

            if state == 'OH':
                template_filename = re.sub(
                    r'\(.+?\)\.docx',
                    f'({ "配偶者有" if fiance else "独身"}).docx',
                    template_filename,
                    count=1)

            template_filename = before_ext.sub(
                f'_{state}',
                template_filename,
                count=1)

            if form_no == 21:
                template_filename = before_ext.sub(
                    f'{"_Chacot"}',
                    template_filename,
                    count=1)
        # NOTEに対しての特殊処理
        elif form_no in form_no_note:
            if state == 'CA':
                template_filename = re.sub(
                    r'\(.+?\)\.docx',
                    '(アモチ有無).docx',
                    template_filename)

            template_filename = before_ext.sub(
                f'_{state}',
                template_filename)

            if form_no == 20:
                template_filename = before_ext.sub(
                    f'{"_Chacot"}',
                    template_filename,
                    count=1)
            return template_filename

        return template_filename

    def get_kv_for_product(self, product_name: str, form_no: int, product_state: str) -> Dict[str, Any]:
        keywords = {k: v for k, v in itertools.chain(
            ((record.key, record.value) for record in self.__keywords_rep_by_product
             if record.name == product_name and record.form_no == form_no),
            ((record.key, record.value) for record in self.__keywords_rep_by_state
             if record.form_no == form_no and record.state == product_state),
            self.__global
        )}

        return keywords

    def get_product_doc_info(self, product: ProductInput) -> Mapping[int, str]:
        """
        商品情報から様式区分とテンプレート名の組を返す
        """
        value = {form_no: self.__specialize_filename(file_name, form_no, product.state, product.fiance is not None)
                 for product_name, form_no, _, file_name in self.__mapping
                 if product_name == product.name
                 if file_name is not None
                 }

        return value

    def get_template_filename(self, product: Product, form_no: int) -> str:
        _, _, _, template_filename = first_true(
            self.__mapping,
            default=(None, None, None, None),
            pred=lambda record: record[0] == product.name and record[1] == form_no,
        )

        if template_filename is None:
            raise RuntimeError(
                f'商品: {product.name}, 様式番号{form_no} が設定情報から見つかりません')

        return self.__specialize_filename(template_filename, form_no, product.state, product.fiance is not None)

    def get_jikkin_sheet_form_no(self, product_name: str) -> int:

        _, form_no, *_ = more_itertools.first_true(
            self.__mapping,
            (None,) * 4,
            pred=lambda tp: all((
                tp[0] == product_name,
                tp[1] in range(7, 10),
                tp[3] is None))
        )
        if form_no == None:
            raise RuntimeError(
                f'{product_name}に該当する実金シートの様式区分が存在しません。設定情報を確認してください')

        return form_no


def load_config_from_workbook(config_xlsx_path: str) -> Config:
    """
    設定情報.xlxsをもとに設定情報オブジェクトを構築する
    """

    wb = load_workbook(config_xlsx_path, data_only=True)
    output_mapping_sheet = cast(ReadOnlyWorksheet, wb['出力様式一覧'])
    sheet_replace_by_state = cast(ReadOnlyWorksheet, wb["州による文章差し替え"])
    sheeet_replace_by_product = cast(ReadOnlyWorksheet, wb['商品による文書差し替え'])
    sheeet_global = cast(ReadOnlyWorksheet, wb['マスタ'])

    def update_at(ls: List[Any], idx: int, f: Callable[[Any], Any]) -> List[Any]:
        ls[idx] = f(ls[idx])
        return ls

    # 商品区分, 様式番号, 様式名, テンプレートファイル名
    mapping = [
        tuple(update_at([cell_normalize(cell) for cell in row], 0, str))
        for row in output_mapping_sheet[f'B2:E{output_mapping_sheet.max_row}']
        if any(cell.value is not None for cell in row)
    ]

    # 様式番号, 様式名, 項目名, アメリカ州, 内容
    replace_by_states = [
        RepByStateRecord(*(cell_normalize(cell) for cell in row))
        for row in sheet_replace_by_state[f'B2:F{sheet_replace_by_state.max_row}']
        if any(cell.value is not None for cell in row)
    ]

    # 商品区分, 様式番号, 様式名, 項目名, 内容
    replace_by_product = [
        RepByProductRecord(
            *update_at([cell_normalize(cell) for cell in row], 0, str))
        for row in sheeet_replace_by_product[f'B2:F{sheeet_replace_by_product.max_row}']
        if any(cell.value is not None for cell in row)
    ]

    # 項目名, 内容
    global_mapping = [tuple(map(cell_normalize, row))
                      for row in sheeet_global[f'B2:C{sheeet_global.max_row}']]

    return Config(mapping, replace_by_states, replace_by_product, global_mapping)
