"""
実金シートの生成処理をまとめたモジュール
"""

from app import config
from app import replace
from app import xl_helper
from app.model import Product, ProductInput, JikkinKV, TableKV
from dataclasses import dataclass
from logging import getLogger
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, ChainMap, Dict, Iterable, Tuple, cast
import openpyxl
import os
import platform

logger = getLogger(__name__)


def range_to_kv(r: Iterable[Tuple[Cell, ...]]) -> Dict[str, Any]:
    """
    選択範囲をdictにして返す
    """
    return {k.value.rstrip(): v.value for k, v in r if isinstance(k.value, str)}


@dataclass(frozen=True)
class ProductBuilder:
    product_input: ProductInput
    src_jikkin_path: str

    def run(self) -> Product:

        stat = os.stat(self.src_jikkin_path)
        if platform.system() == 'Windows':
            if stat.st_ctime == stat.st_mtime:
                raise RuntimeError(
                    f"{self.src_jikkin_path}をExcelから開き、同じ名前で名前をつけて保存を実行してください")
        # macでの処理
        elif platform.system() == 'Darwin':
            if stat.st_birthtime == stat.st_mtime:
                raise RuntimeError(
                    f"{self.src_jikkin_path}をExcelから開き、同じ名前で名前をつけて保存を実行してください")

        p_kv, t_kv = jikkin_to_kv(self.src_jikkin_path)

        return Product(self.product_input, p_kv, t_kv, self.src_jikkin_path)


def output_jikkin_sheet(product_input: ProductInput, template_path: str, output_path: str, config: config.Config) -> ProductBuilder:

    wb = openpyxl.load_workbook(template_path)
    for k, v in zip(*wb['入力シート'].iter_cols(0, 2)):
        if k.value is None:
            continue
        v.value = product_input.get(k.value)

    write_table_to(wb)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    product_name = product_input.product_kv['商品区分']
    product_state = product_input.product_kv['州国']

    logger.info(f'get jikkin sheet {config.get_jikkin_sheet_form_no}')
    replace.replace(output_path, output_path,
                    ChainMap(
                        product_input.product_kv,
                        config.get_kv_for_product(
                            product_name,
                            config.get_jikkin_sheet_form_no(product_name),
                            product_state
                        )
                    ))

    return ProductBuilder(product_input, output_path)


def jikkin_type(product_name: str) -> str:
    # FIXME: 辞書にすることを検討
    if product_name in {'70N', '70NP', 'PM70N', 'PM70NP', 'マルチ50', 'Chacot', 'マルチChacot'}:
        return 'A'
    elif product_name in {'コバルト70', 'マルチ70', 'コバルト60', 'コバルト80'}:
        return 'B'
    elif product_name in {'9054', '2054'}:
        return 'C'
    elif product_name in {'コーポレート50', 'コーポレートマルチ50', '資産管理法人70N', '資産管理法人70NP', '資産管理法人マルチ50', '資産管理法人Chacot', '資産管理法人マルチChacot','コーポレートChacot','コーポレートマルチChacot'}:
        return 'D'
    elif product_name in {'コーポレート60', 'コーポレート70', 'コーポレートマルチ70', '資産管理法人コバルト70', '資産管理法人マルチ70', '資産管理法人コバルト60', '資産管理法人コバルト80'}:
        return 'E'
    else:
        raise RuntimeError(f'{product_name}は本プログラムで処理することができません。')


def get_keywords_for_each_products(ws: Worksheet) -> Iterable[ProductInput]:
    """
    入力シートからMapping[キーワード, 値]のようなものを配列にして返す。

    ```
    get_keywords_for_each_products(wb['入力シート'])[0] # 最初の物件
    get_keywords_for_each_products(wb['入力シート'])[1] # ２つ目の物件
    ```
    """

    columns = ws.iter_cols()  # 先頭がキー、残りが値のリスト
    keys = [cell.value.rstrip() for cell in
            next(columns)  # 先頭のキーだけ取り出し、columnsの中身をvaluesにする
            ]
    values_for_each_product = list(columns)  # 商品ごとの値たち

    def normalize(v: Any):
        if isinstance(v, str):
            return None if len(v := v.rstrip()) == 0 else v
        else:
            return v

    # NOTE: 入力される商品情報の各列は、値が存在しない場合は直前の行の値を取得するという
    # 処理を再帰的に行うようにする仕様である。そのためChainMapの性質を利用し、キーが存在しない
    # 場合に次の辞書を読みに行く機構を利用することにした。最後に参照することになる最終的な値と
    # して1行目の値をkey => Noneを許す辞書を作成し、各行ごとにkey => Noneを許さない形で
    # 単方向連結リストの構造でChainMapをネストさせる。これによって値がない列は直前の値を参照
    # するという仕様が満たせる。

    acc = {k: normalize(cell.value) for k, cell
           in zip(keys, values_for_each_product[0])}

    for values in values_for_each_product:
        record = {k: v for k, cell in zip(keys, values)
                  if (v := normalize(cell.value)) is not None}
        if record.get('商品区分', '').strip() == '':
            continue
        acc = ChainMap(record, acc)
        r = ProductInput(acc)
        yield r


def product_input(wb: Workbook) -> JikkinKV:
    """
    実金.xlsxの入力シートからkvを生成する。入力シート.xlsxに対しての処理でないことに注意
    """

    input_sheet = cast(Worksheet, wb['入力シート'])
    return range_to_kv(zip(*input_sheet.iter_cols(1, 2)))


def write_table_to(wb: Workbook):
    product_name = product_input(wb).get('商品区分')

    t = jikkin_type((product_name))
    if t == 'A':
        return process_a_d(wb)
    if t == 'B':
        return process_b(wb)
    if t == 'C':
        return process_c(wb)
    if t == 'D':
        return process_a_d(wb)
    if t == 'E':
        return process_e(wb)


def process_a_d(workbook: Workbook):
    jikkin_sheet = cast(Worksheet, workbook['実金'])
    kv = product_input(workbook)

    rows = {
        "毎月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            '',
            f'=INT(F{idx-1} * $B$13 / 12)',
            0,
            f'=F{idx-1}-E{idx}',
            f'=IF(F{idx}>0, F{idx}, 0)',
            f'=IF(F{idx-1}>0, ROUNDDOWN(G{idx-1} * I{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")'
        ),
        "最終月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            '',
            f'=INT(F{idx-1} * $B$13 / 12)',
            f'=ABS(F{idx-1})',
            0,
            0,
            f'=IF(F{idx-1}>0, ROUNDDOWN(G{idx-1} * I{idx} / 365, 0))',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")'
        ),
        "合計": lambda start, end: (
            '合計',
            '',
            f'=SUM(C{start}:C{end})',
            f'=SUM(D{start}:D{end})',
            f'=SUM(E{start}:E{end})',
            '',
            '',
            f'=SUM(H{start}:H{end})',
            f'=SUM(I{start}:I{end})',
        )
    }

    repeat_start_idx = 23
    repeat = kv['弁済回数']
    repeat_end_idx = repeat_start_idx + repeat - 3  # 二回目から最終回直前までなので、-3
    min_col = 1
    max_col = 9  # I列
    table = list(jikkin_sheet.iter_rows(
        min_row=repeat_start_idx,
        max_row=repeat_end_idx + 2,
        min_col=min_col,
        max_col=max_col))

    style_row = next(jikkin_sheet.iter_rows(
        repeat_start_idx-1, repeat_start_idx-1))

    for row in table:
        row_idx = row[0].row
        if repeat_start_idx <= row_idx <= repeat_end_idx:
            src_row = rows['毎月'](row_idx)
        elif row_idx == repeat_end_idx + 1:
            src_row = rows['最終月'](row_idx)
        elif row_idx == repeat_end_idx + 2:
            src_row = rows['合計'](row_idx - 1 - repeat, row_idx - 1)

        assert (len(src_row) == max_col - min_col + 1)
        for cell, v in zip(row, src_row):
            cell.value = v

        # スタイルの適用
        for s, d in zip(style_row, row):
            xl_helper.copy_style(s, d)

    # 最終月や合計に依存するセルを更新

    last_month_idx = repeat_end_idx + 1
    sum_idx = last_month_idx + 1

    goldPatternFill = PatternFill(patternType='solid', fgColor='fff2cc')
    jikkin_sheet[f'E{last_month_idx}'].fill = goldPatternFill
    jikkin_sheet[f'H{sum_idx}'].fill = goldPatternFill

    jikkin_sheet['E6'].value = f'=ROUNDDOWN(G6/H{sum_idx}, 5)'
    jikkin_sheet['E7'].value = f'=D{sum_idx} + E{sum_idx}'

    jikkin_sheet['E10'].value = f'=D{last_month_idx} + E{last_month_idx}'
    jikkin_sheet['G6'].value = f'=C{sum_idx} + D{sum_idx}'
    jikkin_sheet['G7'].value = f'=ABS(E{sum_idx})'


def process_b(workbook: Workbook):
    """
    端数処理ありバルーン
    """
    input_sheet = cast(Worksheet, workbook['入力シート'])
    jikkin_sheet = cast(Worksheet, workbook['実金'])

    kv = product_input(workbook)

    rows = {
        "毎月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=IF(C{idx}<>"", $H$10, 0)',
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$22, A{idx}<={input_sheet.title}!$B$40), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
        ),
        "最終月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=G{idx-1}',  # ここだけちがう
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$22, A{idx}<={input_sheet.title}!$B$40), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
        ),
        "合計": lambda start, end: (
            '合計',
            '',
            '',
            f'=SUM(D{start}:D{end})',
            f'=SUM(E{start}:E{end})',
            f'=SUM(F{start}:F{end})',
            f'=SUM(G{start}:G{end})',
            f'=SUM(H{start}:H{end})',
            f'=SUM(I{start}:I{end})',
            f'=SUM(J{start}:J{end})',
            '',  # -------------------------------------------------------------
            '',
            '',
        )
    }

    repeat_start_idx = 24
    repeat = kv['弁済回数']
    repeat_end_idx = repeat_start_idx + repeat - 3  # 二回目から最終回直前までなので、-3
    min_col = 1
    max_col = 13  # M列(13)
    table = list(jikkin_sheet.iter_rows(
        min_row=repeat_start_idx,
        max_row=repeat_end_idx + 2,
        min_col=min_col,
        max_col=max_col))

    style_row = next(jikkin_sheet.iter_rows(
        repeat_start_idx-1, repeat_start_idx-1))

    for row in table:
        row_idx = row[0].row
        if repeat_start_idx <= row_idx <= repeat_end_idx:
            src_row = rows['毎月'](row_idx)
        elif row_idx == repeat_end_idx + 1:
            src_row = rows['最終月'](row_idx)
        elif row_idx == repeat_end_idx + 2:
            src_row = rows['合計'](row_idx - 1 - repeat, row_idx - 1)

        assert (len(src_row) == max_col - min_col + 1)
        for cell, v in zip(row, src_row):
            cell.value = v

        # スタイルの適用
        for s, d in zip(style_row, row):
            xl_helper.copy_style(s, d)

    # 最終月や合計に依存するセルを更新

    last_month_idx = repeat_end_idx + 1
    sum_idx = last_month_idx + 1

    jikkin_sheet['F7'].value = f'=ROUNDDOWN($H$7/$I${sum_idx}, 5)'
    jikkin_sheet['F8'].value = f'=$E${sum_idx} + $F${sum_idx}'
    jikkin_sheet['F11'].value = f'=$E${last_month_idx} + $F${last_month_idx}'
    jikkin_sheet['F12'].value = f'=VLOOKUP(入力シート!$B$22,$B$22:$F{last_month_idx},5,0)'

    jikkin_sheet['H5'].value = f'=F{last_month_idx}/($B$4 * $B$7)'
    jikkin_sheet['H7'].value = f'=$D${sum_idx} + $E${sum_idx}'
    jikkin_sheet['H8'].value = f'=ABS($F${sum_idx})'


def process_c(workbook: Workbook):
    """
    端数処理ありバルーン
    """
    input_sheet = cast(Worksheet, workbook['入力シート'])
    jikkin_sheet = cast(Worksheet, workbook['実金'])
    kv = product_input(workbook)

    rows = {
        "毎月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=IF(C{idx}<>"", $H$10, 0)',
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$22, A{idx}<={input_sheet.title}!$B$40), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
            f'=IF(N$22>=M{idx}, 1, 0)'
        ),
        "最終月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=G{idx-1}',  # ここだけちがう
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$22, A{idx}<={input_sheet.title}!$B$40), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
            f'=IF(N$22>=M{idx}, 1, 0)'
        ),
        "合計": lambda start, end: (
            '合計',
            '',
            '',
            f'=SUM(D{start}:D{end})',
            f'=SUM(E{start}:E{end})',
            f'=SUM(F{start}:F{end})',
            '',
            '', 
            f'=SUM(I{start}:I{end})',
            f'=SUM(J{start}:J{end})',
            '',  # -------------------------------------------------------------
            '',
            '',
            '',
        )
    }

    repeat_start_idx = 24
    repeat = kv['弁済回数']
    repeat_end_idx = repeat_start_idx + repeat - 3  # 二回目から最終回直前までなので、-3
    min_col = 1
    max_col = 14  # N列(14)
    table = list(jikkin_sheet.iter_rows(
        min_row=repeat_start_idx,
        max_row=repeat_end_idx + 2,
        min_col=min_col,
        max_col=max_col))

    style_row = next(jikkin_sheet.iter_rows(
        repeat_start_idx-1, repeat_start_idx-1))
    for row in table:
        row_idx = row[0].row
        if repeat_start_idx <= row_idx <= repeat_end_idx:
            src_row = rows['毎月'](row_idx)
        elif row_idx == repeat_end_idx + 1:
            src_row = rows['最終月'](row_idx)
        elif row_idx == repeat_end_idx + 2:
            src_row = rows['合計'](row_idx - 1 - repeat, row_idx - 1)

        assert (len(src_row) == max_col - min_col + 1)
        for cell, v in zip(row, src_row):
            cell.value = v

        # スタイルの適用
        for s, d in zip(style_row, row):
            xl_helper.copy_style(s, d)

    # 最終月や合計に依存するセルを更新

    last_month_idx = repeat_end_idx + 1
    sum_idx = last_month_idx + 1

    jikkin_sheet['F7'].value = f'=ROUNDDOWN($H$7/$I${sum_idx}, 5)'
    jikkin_sheet['F8'].value = f'=$E${sum_idx} + $F${sum_idx}'
    jikkin_sheet['F11'].value = f'=$E${last_month_idx} + $F${last_month_idx}'
    jikkin_sheet['F12'].value = f'=VLOOKUP({input_sheet.title}!$B$22,$B$22:$F{last_month_idx},5,0)'

    jikkin_sheet['H5'].value = f'=F{sum_idx}/($B$4 * $B$7)'
    jikkin_sheet['H7'].value = f'=$D${sum_idx} + $E${sum_idx}'
    jikkin_sheet['H8'].value = f'=ABS($F${sum_idx})'


def process_e(workbook: Workbook):
    """
    端数処理ありバルーン
    """
    input_sheet = cast(Worksheet, workbook['入力シート'])
    jikkin_sheet = cast(Worksheet, workbook['実金'])

    kv = product_input(workbook)

    rows = {
        "毎月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=IF(C{idx}<>"", $H$10, 0)',
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$24, A{idx}<={input_sheet.title}!$B$42), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
        ),
        "最終月": lambda idx: (
            f'=A{idx-1} + 1',
            f'=EDATE(B{idx-1}, 1)',
            f'=IF(AND(L{idx}=1,$F$4>=M{idx}), M{idx}, "")',
            '',
            f'=INT(G{idx-1} * $B$14 / 12)',
            f'=G{idx-1}',  # ここだけちがう
            f'=G{idx-1}-F{idx}',
            f'=IF(G{idx}>0, G{idx}, 0)',
            f'=IF(G{idx-1}>0, ROUNDDOWN(H{idx-1} * J{idx} / 365, 0), 0)',
            f'=DATEDIF(B{idx-1}, B{idx}, "d")',
            '',  # -------------------------------------------------------------
            f'=IF(AND(B{idx}>={input_sheet.title}!$B$24, A{idx}<={input_sheet.title}!$B$42), IF(MONTH(B{idx})=F$5, 1, 0), 0)',
            f'=SUM(L$22:L{idx})',
        ),
        "合計": lambda start, end: (
            '合計',
            '',
            '',
            f'=SUM(D{start}:D{end})',
            f'=SUM(E{start}:E{end})',
            f'=SUM(F{start}:F{end})',
            f'=SUM(G{start}:G{end})',
            f'=SUM(H{start}:H{end})',
            f'=SUM(I{start}:I{end})',
            f'=SUM(J{start}:J{end})',
            '',  # -------------------------------------------------------------
            '',
            '',
        )
    }

    repeat_start_idx = 24
    repeat = kv['弁済回数']
    repeat_end_idx = repeat_start_idx + repeat - 3  # 二回目から最終回直前までなので、-3
    min_col = 1
    max_col = 13  # M列(13)
    table = list(jikkin_sheet.iter_rows(
        min_row=repeat_start_idx,
        max_row=repeat_end_idx + 2,
        min_col=min_col,
        max_col=max_col))

    style_row = next(jikkin_sheet.iter_rows(
        repeat_start_idx-1, repeat_start_idx-1))

    for row in table:
        row_idx = row[0].row
        if repeat_start_idx <= row_idx <= repeat_end_idx:
            src_row = rows['毎月'](row_idx)
        elif row_idx == repeat_end_idx + 1:
            src_row = rows['最終月'](row_idx)
        elif row_idx == repeat_end_idx + 2:
            src_row = rows['合計'](row_idx - 1 - repeat, row_idx - 1)

        assert (len(src_row) == max_col - min_col + 1)
        for cell, v in zip(row, src_row):
            cell.value = v

        # スタイルの適用
        for s, d in zip(style_row, row):
            xl_helper.copy_style(s, d)

    # 最終月や合計に依存するセルを更新

    last_month_idx = repeat_end_idx + 1
    sum_idx = last_month_idx + 1

    jikkin_sheet['F7'].value = f'=ROUNDDOWN($H$7/$I${sum_idx}, 5)'
    jikkin_sheet['F8'].value = f'=$E${sum_idx} + $F${sum_idx}'
    jikkin_sheet['F11'].value = f'=$E${last_month_idx} + $F${last_month_idx}'
    jikkin_sheet['F12'].value = f'=VLOOKUP({input_sheet.title}!$B$24,$B$22:$F{last_month_idx},5,0)'

    jikkin_sheet['H5'].value = f'=F{last_month_idx}/($B$4 * $B$7)'
    jikkin_sheet['H7'].value = f'=$D${sum_idx} + $E${sum_idx}'
    jikkin_sheet['H8'].value = f'=ABS($F${sum_idx})'


def jikkin_to_kv(jikkin_path: str) -> Tuple[JikkinKV, TableKV]:
    """テーブルのシートからkvを生成する"""
    wb = openpyxl.load_workbook(jikkin_path, data_only=True)
    return (product_input(wb), table_kv(wb))


def table_kv(wb: Workbook) -> Dict[str, Any]:
    p = product_input(wb)
    t = jikkin_type(p['商品区分'])

    table_sheet = cast(Worksheet, wb['実金'])

    if t in {'A', 'D'}:
        return {
            **range_to_kv(table_sheet['A3:B15']),
            **range_to_kv(table_sheet['D6:E10']),
            **range_to_kv(table_sheet['C15:D16']),
            **range_to_kv(zip(*table_sheet['F15:H16'])),
            **range_to_kv(zip(*table_sheet['H17:H18'])),
        }
    else:
        return {
            **range_to_kv(table_sheet['A3:B16']),
            **range_to_kv(table_sheet['E3:F12']),
            **range_to_kv(table_sheet['G4:H12']),
            **range_to_kv(table_sheet['D16:E17']),
            **range_to_kv(zip(*table_sheet['G16:I17'])),
            **range_to_kv(zip(*table_sheet['I16:I17'])),
        }
