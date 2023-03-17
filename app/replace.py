"""
キーワード置換をするための知識をもたせたモジュール
"""

from app import docx_helper, time_helper, xl_helper
from dataclasses import dataclass
from logging import getLogger
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Mapping, cast
import docx
import itertools
import mojimoji
import openpyxl
import os
import re
import datetime

# wordで置換する文字列を囲う文字
keyword_quoter = '●'
format_separator = '○'
replace_pattern = re.compile(
    rf"{keyword_quoter}(.+?)(?:{format_separator}(.+?))?{keyword_quoter}")

logger = getLogger(__name__)


@dataclass(frozen=True)
class TemplateReplaceInfo:
    """
    テンプレート置換をするのに必要な情報
    """
    form_no: str  # 様式番号
    template_file_name: str  # テンプレートファイル名
    output_file_name: str  # 出力ファイル名
    keywords: Mapping[str, Any]  # 置換キーワード


def replace_template_string(input_string: str, replace_dict: Mapping[str, Any], logging_input_file_path: str):
    """
    strに含まれているテンプレート文字列"●項目名○フォーマット●"のようなものを置換する。
    """
    output_string = input_string

    matched = replace_pattern.search(input_string)
    if matched is None:
        return output_string
    
    # プログラムの意味的にはwhileによるループでよいが、無限に回ってしまうのは問題があるので
    # for文で数えながら置換を行い、20回以上置換が発生する場合はエラーを投げて終了する
    for c in itertools.count():

        matched = replace_pattern.search(output_string)
        if matched is None:
            break

        if c >= 20:
            logger.debug(output_string)
            logger.debug(f'{matched}')
            logger.debug(f'{replace_pattern.search(output_string)}')
            raise RuntimeError('テキストの置換に失敗しました。')

        key, format_str = matched.groups()
        value = replace_dict.get(key)

        if value is None:
            logger.debug(f'{logging_input_file_path}に含まれる、{key!r}は置換できません')
            value = ''
        # フォーマット適用
        if isinstance(value, (datetime.date, datetime.time)):
            value = time_helper.strftime(
                value,
                mojimoji.zen_to_han(format_str)
                if format_str is not None else "%Y年%-m月%-d日").upper()
        elif isinstance(value, (int, float)):
            format_str = f"{{:{mojimoji.zen_to_han(format_str)}}}" \
                if format_str is not None else '{:,}'
            value = format_str.format(value)

        output_string = replace_pattern.sub(value, output_string, 1)

    return output_string


def replace(input_file_path: str, output_file_path: str, replace_dict: Mapping[str, Any]):
    """
    replace_dictをもとにinput_file_pathのドキュメントに含まれるキーワードを置換し、output_file_pathに出力する
    対応フォーマットはxlsxおよびdocx
    """

    if re.search(r'.xlsx$', input_file_path):
        return replace_excel(input_file_path, output_file_path, replace_dict)
    else:
        return replace_word(input_file_path, output_file_path, replace_dict)


def replace_word(input_file_path: str, output_file_path: str, replace_dict: Mapping[str, Any]):
    """
    wordの置換を行う関数
    特に理由がなければこちらの関数を直接呼び出さず、`replace`を使うべき
    """
    doc = docx.Document(docx=input_file_path)
    for run in docx_helper.all_runs(doc):
        if isinstance(run.text, str):
            replaced_str = replace_template_string(
                run.text, replace_dict, input_file_path)
            if run.text is not replaced_str:
                run.text = replaced_str
    doc.save(output_file_path)


def replace_excel(input_file_path: str, output_file_path: str, replace_dict: Mapping[str, Any]):
    """
    excelの置換を行う関数
    特に理由がなければこちらの関数を直接呼び出さず、`replace`を使うべき
    """

    wb = openpyxl.load_workbook(input_file_path)

    for ws in wb:
        ws = cast(Worksheet, ws)

        # ヘッダーとフッターの置換
        for headerPart in xl_helper.all_header_footer_parts(ws):
            if not isinstance(headerPart.text, str):
                continue
            replace_txt = replace_template_string(
                headerPart.text, replace_dict, input_file_path)
            if headerPart.text is not replace_txt:
                headerPart.text = replace_txt

        # セル内の文字列置換
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row,
                                min_col=ws.min_column, max_col=ws.max_column):
            for cell in row:
                if (isinstance(cell.value, str)):
                    replaced_str = replace_template_string(
                        cell.value, replace_dict, input_file_path)
                    if cell.value is not replaced_str:
                        cell.value = replaced_str

    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    wb.save(output_file_path)
